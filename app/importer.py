"""Excel 匯入：目前支援 fig1（新竹採購進貨）。"""
from __future__ import annotations
from datetime import datetime, date
from typing import Optional
from openpyxl import load_workbook

from . import db


FIG1_HEADERS = [
    "到貨日期", "簽收人", "供應商", "品牌", "產品名稱",
    "序號", "進貨數量", "請購人員", "請購PO", "存放位置",
]
# 供應商欄位的可接受別名（規範名為「供應商」）
SUPPLIER_ALIASES = ("供應商", "出貨對象")
# 產品名稱欄位的別名
MODEL_ALIASES = ("產品名稱", "料件", "型號")
# 對應工號的別名
PROJECT_ALIASES = ("對應工號", "工號")


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_qty(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        # 容忍像 '8PC'、'10 個'、'5pcs' 這類帶單位的字串：取開頭數字
        import re
        m = re.match(r"\s*(-?\d+(?:\.\d+)?)", str(v))
        return float(m.group(1)) if m else 0.0


def _parse_serials(v) -> list[str]:
    s = _norm(v)
    if not s:
        return []
    out = []
    for token in s.replace(",", "\n").replace("，", "\n").splitlines():
        token = token.strip()
        if token:
            out.append(token)
    return out


def _parse_serials_office(v) -> list[str]:
    """辦公室請購：序號可能以 `/`、`,` 分隔多筆，整段可有 `SN:` 或 `S/N:` 前綴；
    輸出每筆都統一以 `SN:` 開頭。"""
    s = _norm(v)
    if not s:
        return []
    head = s.lstrip()
    for prefix in ("S/N:", "S/N：", "SN:", "SN："):
        if head.upper().startswith(prefix.upper()):
            s = head[len(prefix):]
            break
    out = []
    for tok in s.replace("/", "\n").replace(",", "\n").replace("，", "\n").splitlines():
        tok = tok.strip()
        if not tok:
            continue
        for prefix in ("S/N:", "S/N：", "SN:", "SN："):
            if tok.upper().startswith(prefix.upper()):
                tok = tok[len(prefix):].strip()
                break
        if tok:
            out.append(f"SN:{tok}")
    return out


def _get_or_create(c, table: str, key_col: str, key_val: str,
                   extra: dict | None = None) -> int:
    row = c.execute(f"SELECT id FROM {table} WHERE {key_col}=?", (key_val,)).fetchone()
    if row:
        return row["id"]
    cols = [key_col] + list((extra or {}).keys())
    vals = [key_val] + list((extra or {}).values())
    placeholders = ",".join(["?"] * len(cols))
    cur = c.execute(f"INSERT INTO {table}({','.join(cols)}) VALUES({placeholders})", vals)
    return cur.lastrowid


def _resolve_product_for_inbound(c, brand_label: str, model_label: str,
                                  stats: dict, row_no: int):
    """回傳 (product_row or None, brand_id or None, ok)。
    brand 有填：(brand_id, model) 查找；找不到後續會以 brand 建立新料件。
    brand 空白：以 model 單獨查；找到唯一一筆→用它；多筆衝突→錯誤；找不到→錯誤
                （沒品牌時不替使用者建料件，避免產生空品牌）。
    """
    if brand_label:
        brand_id = _get_or_create(c, "brands", "name", brand_label)
        row = c.execute("SELECT id, is_kit FROM products WHERE brand_id=? AND model=?",
                        (brand_id, model_label)).fetchone()
        return row, brand_id, True
    # 沒品牌 — 以 model 查
    matches = c.execute("SELECT id, is_kit, brand_id FROM products WHERE model=?",
                        (model_label,)).fetchall()
    if len(matches) == 1:
        return matches[0], matches[0]["brand_id"], True
    if len(matches) == 0:
        stats["errors"].append({"row": row_no,
            "msgs": [f"找不到料件「{model_label}」，且未提供品牌可新建"]})
        return None, None, False
    stats["errors"].append({"row": row_no,
        "msgs": [f"料件「{model_label}」在主檔有 {len(matches)} 筆同名（請在 Excel 補上品牌以區分）"]})
    return None, None, False


def _expand_kit_or_insert_line(c, in_id: int, brand_label: str, model_label: str,
                               product_row, qty: float, loc_id, serials: list[str],
                               stats: dict, row_no: int, project_id=None,
                               supplier_id=None) -> int:
    """寫一筆 inbound_line。回傳寫入的明細數。
    product_row: {id} 或 None（None 表示之後需新建）。
    """
    if product_row:
        prod_id = product_row["id"]
    else:
        prod_id = _get_or_create_product(c, _get_or_create(c, "brands", "name", brand_label),
                                          model_label, "")
    cur2 = c.execute("""INSERT INTO inbound_lines(inbound_id, product_id, qty, unit,
                        location_id, is_surplus, project_id, supplier_id)
                        VALUES(?,?,?,?,?,0,?,?)""",
                     (in_id, prod_id, qty, "個", loc_id, project_id, supplier_id))
    line_id = cur2.lastrowid
    for sn in serials:
        c.execute("""INSERT OR IGNORE INTO serial_items(product_id, serial_no, status,
                     current_location_id, inbound_line_id, is_surplus, project_id)
                     VALUES(?,?,?,?,?,0,?)""",
                  (prod_id, sn, "in_stock", loc_id, line_id, project_id))
    return 1


def _get_or_create_product(c, brand_id: int, model: str, description: str) -> int:
    row = c.execute("SELECT id, description FROM products WHERE brand_id=? AND model=?",
                    (brand_id, model)).fetchone()
    if row:
        # 若原本沒有敘述但匯入有，補上
        if description and not row["description"]:
            c.execute("UPDATE products SET description=? WHERE id=?", (description, row["id"]))
        return row["id"]
    cur = c.execute("""INSERT INTO products(brand_id, model, description, base_unit, track_by_serial)
                       VALUES(?,?,?,?,0)""", (brand_id, model, description or None, "個"))
    return cur.lastrowid


PARTS_MODEL_ALIASES = ("料件", "型號", "產品名稱", "新增料件")
PARTS_DESC_ALIASES = ("敘述", "產品敘述", "描述", "簽收人")  # 簽收人為舊檔誤植，沿用相容


def build_parts_template() -> bytes:
    """產生一份只有表頭 + 一筆示範資料的料件主檔範本。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "料件"
    ws.append(["品牌", "料件", "敘述"])
    hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hb; cell.alignment = Alignment(horizontal="center")
    ws.append(["AB", "1756-IB32", "10-31 VDC INPUT 32 PTS (36 PIN)"])
    for col, w in zip("ABC", (12, 22, 50)):
        ws.column_dimensions[col].width = w
    buf = BytesIO(); wb.save(buf); return buf.getvalue()


def import_parts(file_bytes: bytes, dry_run: bool = False) -> dict:
    """匯入料件主檔。強制 is_kit=0。已存在的 (brand, model) 直接跳過。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "valid_rows": 0, "inserted": 0, "skipped_existing": 0,
                "errors": [], "dry_run": dry_run, "details": []}
    header = [_norm(x) for x in rows[0]]

    def find_idx(canonical, aliases):
        if canonical in header:
            return header.index(canonical)
        for a in aliases:
            if a in header:
                return header.index(a)
        return None

    brand_idx = header.index("品牌") if "品牌" in header else None
    model_idx = find_idx("料件", PARTS_MODEL_ALIASES)
    desc_idx = find_idx("敘述", PARTS_DESC_ALIASES)

    missing = []
    if brand_idx is None: missing.append("品牌")
    if model_idx is None: missing.append("料件 (可接受：料件/型號/產品名稱/新增料件)")
    if missing:
        raise ValueError(f"缺少必要欄位：{', '.join(missing)}")

    errors = []
    cleaned = []
    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or _norm(c) == "" for c in raw):
            continue
        brand = _norm(raw[brand_idx]) if brand_idx < len(raw) else ""
        model = _norm(raw[model_idx]) if model_idx < len(raw) else ""
        desc = _norm(raw[desc_idx]) if (desc_idx is not None and desc_idx < len(raw)) else ""
        msgs = []
        if not brand: msgs.append("缺品牌")
        if not model: msgs.append("缺料件")
        if msgs:
            errors.append({"row": i, "msgs": msgs}); continue
        cleaned.append({"row": i, "brand": brand, "model": model, "description": desc})

    stats = {
        "total_rows": len(rows) - 1,
        "valid_rows": len(cleaned),
        "inserted": 0,
        "skipped_existing": 0,
        "errors": errors,
        "dry_run": dry_run,
        "details": [],
    }

    if dry_run:
        for r in cleaned:
            stats["details"].append({**r, "action": "would-insert"})
        return stats

    with db.tx() as c:
        for r in cleaned:
            brand_id = _get_or_create(c, "brands", "name", r["brand"])
            existing = c.execute("SELECT id FROM products WHERE brand_id=? AND model=?",
                                 (brand_id, r["model"])).fetchone()
            if existing:
                stats["skipped_existing"] += 1
                stats["details"].append({**r, "action": "skipped (已存在)"})
                continue
            c.execute("""INSERT INTO products(brand_id, model, description, base_unit,
                         track_by_serial, safety_stock, is_kit)
                         VALUES(?,?,?,?,?,?,0)""",
                      (brand_id, r["model"], r["description"] or None, "個", 0, 0))
            stats["inserted"] += 1
            stats["details"].append({**r, "action": "inserted"})
    return stats


def build_projects_template() -> bytes:
    """工號 / 案件範本。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "工號"
    ws.append(["工號", "業主", "案名"])
    hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hb; cell.alignment = Alignment(horizontal="center")
    ws.append(["J115-05-192", "兆聯實業", "TSMC_F18P9_WWT+REC 系統儀控工程"])
    for col, w in zip("ABC", (18, 16, 60)):
        ws.column_dimensions[col].width = w
    buf = BytesIO(); wb.save(buf); return buf.getvalue()


def import_projects(file_bytes: bytes, dry_run: bool = False) -> dict:
    """匯入工號 / 案件主檔。已存在的工號直接跳過。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "valid_rows": 0, "inserted": 0, "skipped_existing": 0,
                "errors": [], "dry_run": dry_run, "details": []}
    header = [_norm(x) for x in rows[0]]

    job_aliases = ("工號", "工單", "案號", "job_no")
    owner_aliases = ("業主", "客戶", "業主名稱", "owner")
    name_aliases = ("案名", "工程名稱", "專案名稱", "project_name")

    def idx_of(aliases):
        for a in aliases:
            if a in header:
                return header.index(a)
        return None

    job_idx = idx_of(job_aliases)
    owner_idx = idx_of(owner_aliases)
    name_idx = idx_of(name_aliases)

    if job_idx is None:
        raise ValueError("缺少必要欄位：工號（可接受別名：工單/案號）")

    errors = []
    cleaned = []
    seen_in_file = set()
    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or _norm(c) == "" for c in raw):
            continue
        job = _norm(raw[job_idx]) if job_idx < len(raw) else ""
        owner = _norm(raw[owner_idx]) if (owner_idx is not None and owner_idx < len(raw)) else ""
        name = _norm(raw[name_idx]) if (name_idx is not None and name_idx < len(raw)) else ""
        if not job:
            errors.append({"row": i, "msgs": ["缺工號"]}); continue
        if job in seen_in_file:
            errors.append({"row": i, "msgs": [f"工號 {job} 在 Excel 內重複"]}); continue
        seen_in_file.add(job)
        cleaned.append({"row": i, "job_no": job, "owner": owner, "project_name": name})

    stats = {
        "total_rows": len(rows) - 1,
        "valid_rows": len(cleaned),
        "inserted": 0,
        "skipped_existing": 0,
        "errors": errors,
        "dry_run": dry_run,
        "details": [],
    }

    if dry_run:
        for r in cleaned:
            stats["details"].append({**r, "action": "would-insert"})
        return stats

    with db.tx() as c:
        for r in cleaned:
            existing = c.execute("SELECT id FROM projects WHERE job_no=?", (r["job_no"],)).fetchone()
            if existing:
                stats["skipped_existing"] += 1
                stats["details"].append({**r, "action": "skipped (已存在)"})
                continue
            c.execute("INSERT INTO projects(job_no, owner, project_name) VALUES(?,?,?)",
                      (r["job_no"], r["owner"] or None, r["project_name"] or None))
            stats["inserted"] += 1
            stats["details"].append({**r, "action": "inserted"})
    return stats


OFFICE_HEADERS = [
    "到貨日期", "簽收人", "品牌", "產品名稱",
    "序號", "進貨數量", "對應工號", "存放位置",
]


def build_office_template() -> bytes:
    """辦公室請購進貨範本。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "辦公室請購"
    ws.append(["項目"] + OFFICE_HEADERS)
    hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hb; cell.alignment = Alignment(horizontal="center")
    # 兩筆示範資料：同項目編號 = 同一張進貨單
    ws.append([1, "2026-06-25", "杜俊毅", "AB", "1769-IQ32", "", 10,
               "J115-05-192", "倉庫棧板"])
    ws.append([1, "2026-06-25", "杜俊毅", "AB", "1769-OB32", "", 5,
               "J115-05-192", "倉庫棧板"])
    widths = [6, 12, 10, 12, 22, 18, 8, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = BytesIO(); wb.save(buf); return buf.getvalue()


def import_office(file_bytes: bytes, dry_run: bool = False, default_project_id: int = None) -> dict:
    """匯入辦公室請購進貨。同 (date, signer, project) 視為一張單。
    若 Excel 沒有「對應工號」欄或某列留空，會使用 default_project_id（必須提供）。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "valid_rows": 0, "groups": 0, "lines_inserted": 0,
                "groups_inserted": 0, "errors": [], "dry_run": dry_run, "details": []}
    header = [_norm(x) for x in rows[0]]
    idx = {h: header.index(h) if h in header else None for h in OFFICE_HEADERS}
    # 別名解析
    if idx["產品名稱"] is None:
        for alias in MODEL_ALIASES:
            if alias in header:
                idx["產品名稱"] = header.index(alias); break
    if idx["對應工號"] is None:
        for alias in PROJECT_ALIASES:
            if alias in header:
                idx["對應工號"] = header.index(alias); break
    missing = [h for h, v in idx.items() if v is None and h in ("到貨日期", "品牌", "產品名稱", "進貨數量")]
    if missing:
        raise ValueError(f"缺少必要欄位：{', '.join(missing)}")

    errors = []
    parsed = []
    default_project_job_no = None
    if default_project_id:
        with db.tx() as c:
            row = c.execute("SELECT job_no FROM projects WHERE id=?", (default_project_id,)).fetchone()
            default_project_job_no = row["job_no"] if row else None

    # 供應商與請購PO 為選填，office 也記錄 — 找欄位 index（含別名）
    supplier_idx = None
    for alias in SUPPLIER_ALIASES:
        if alias in header:
            supplier_idx = header.index(alias); break
    po_idx = header.index("請購PO") if "請購PO" in header else None
    requester_idx = header.index("請購人員") if "請購人員" in header else None
    item_idx = header.index("項目") if "項目" in header else None

    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or _norm(c) == "" for c in raw):
            continue
        def cell(name):
            j = idx[name]
            return raw[j] if j is not None and j < len(raw) else None
        d = _norm(cell("到貨日期"))
        signer = _norm(cell("簽收人"))
        brand = _norm(cell("品牌"))
        model = _norm(cell("產品名稱"))
        sns = _parse_serials_office(cell("序號"))
        qty = _parse_qty(cell("進貨數量"))
        job_no_in_excel = _norm(cell("對應工號"))
        loc = _norm(cell("存放位置"))
        supplier = _norm(raw[supplier_idx]) if (supplier_idx is not None and supplier_idx < len(raw)) else ""
        po_no = _norm(raw[po_idx]) if (po_idx is not None and po_idx < len(raw)) else ""
        requester = _norm(raw[requester_idx]) if (requester_idx is not None and requester_idx < len(raw)) else ""
        item_no = _norm(raw[item_idx]) if (item_idx is not None and item_idx < len(raw)) else ""
        msgs = []
        if not d: msgs.append("缺到貨日期")
        if not model: msgs.append("缺產品名稱")
        if qty <= 0: msgs.append("數量需 > 0")
        if sns and qty > 0 and len(sns) > int(qty):
            msgs.append(f"序號筆數({len(sns)})多於進貨數量({int(qty)})")
        job_no = job_no_in_excel or default_project_job_no or ""
        if msgs:
            errors.append({"row": i, "msgs": msgs}); continue
        parsed.append({"row_no": i, "date": d, "signer": signer, "brand": brand,
                       "model": model, "serials": sns, "qty": qty,
                       "job_no": job_no, "location": loc,
                       "supplier": supplier, "po_no": po_no, "requester": requester,
                       "item_no": item_no})

    groups = {}  # key -> {header: dict, lines: list}
    for r in parsed:
        if item_idx is not None:
            # 以「項目」為唯一分組鍵；單頭欄位於同 group 內取第一筆非空值
            key = r["item_no"] or f"__row{r['row_no']}__"
        else:
            # 沒「項目」欄位時退回舊規則
            key = (r["date"], r["signer"], r["job_no"], r["supplier"], r["po_no"])
        g = groups.setdefault(key, {"header": {}, "lines": []})
        for k in ("date", "signer", "job_no", "supplier", "po_no", "requester"):
            if not g["header"].get(k) and r.get(k):
                g["header"][k] = r[k]
        g["lines"].append(r)

    stats = {
        "total_rows": len(rows) - 1,
        "valid_rows": len(parsed),
        "groups": len(groups),
        "lines_inserted": 0,
        "groups_inserted": 0,
        "errors": errors,
        "dry_run": dry_run,
        "details": [],
    }

    if dry_run:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            stats["details"].append({
                "date": h.get("date"), "signer": h.get("signer"),
                "job_no": h.get("job_no"), "supplier": h.get("supplier"),
                "po_no": h.get("po_no"),
                "lines": len(lines),
                "preview": [f'{ln["brand"]} {ln["model"]} x{ln["qty"]} → {ln["location"]}' for ln in lines],
            })
        return stats

    with db.tx() as c:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            d = h.get("date"); signer = h.get("signer"); job_no = h.get("job_no")
            po_no = h.get("po_no")
            signer_id = _get_or_create(c, "staff", "name", signer, {"role": "簽收"}) if signer else None
            # 供應商：每行可不同。先彙整 distinct 供應商；單頭 supplier_id 取第一個、extra_suppliers 列全部（顯示相容）
            line_supplier_ids = {}  # name -> id, preserve insertion order
            for ln in lines:
                raw_sup = (ln.get("supplier") or "").strip()
                if not raw_sup:
                    continue
                # 同一儲存格內仍可能多家用 '/' / ',' 分隔
                for part in [p.strip() for raw in str(raw_sup).replace(",", "\n").replace("，", "\n").replace("/", "\n").splitlines() for p in [raw] if p.strip()]:
                    if part not in line_supplier_ids:
                        line_supplier_ids[part] = _get_or_create(c, "suppliers", "name", part)
            supplier_id = next(iter(line_supplier_ids.values()), None)
            extra_suppliers = "\n".join(line_supplier_ids.keys()) if len(line_supplier_ids) > 1 else None
            requester_name = next((ln["requester"] for ln in lines if ln.get("requester")), "")
            requester_id = _get_or_create(c, "staff", "name", requester_name,
                                          {"role": "請購"}) if requester_name else None
            po_id = None
            if po_no:
                po_row = c.execute("SELECT id FROM purchase_orders WHERE po_no=?", (po_no,)).fetchone()
                if po_row:
                    po_id = po_row["id"]
                    if requester_id:
                        c.execute("UPDATE purchase_orders SET requester_id=COALESCE(requester_id, ?) WHERE id=?",
                                  (requester_id, po_id))
                else:
                    cur = c.execute("INSERT INTO purchase_orders(po_no, date, requester_id) VALUES(?,?,?)",
                                    (po_no, d, requester_id))
                    po_id = cur.lastrowid
            project_id = None
            extra_job_nos = None
            if job_no:
                # 工號可能含多組，以 '/' 或 ',' 分隔
                parts = [p.strip() for raw in str(job_no).replace(",", "/").replace("，", "/").split("/")
                         for p in [raw] if p.strip()]
                primary = parts[0]
                prow = c.execute("SELECT id FROM projects WHERE job_no=?", (primary,)).fetchone()
                if not prow:
                    cur = c.execute("INSERT INTO projects(job_no) VALUES(?)", (primary,))
                    project_id = cur.lastrowid
                else:
                    project_id = prow["id"]
                if len(parts) > 1:
                    # 全部工號（含 primary）換行存進 extra_job_nos，display 時取代 job_no
                    extra_job_nos = "\n".join(parts)
            cur = c.execute("""INSERT INTO inbound_orders(type, date, signer_id, project_id,
                                                          supplier_id, po_id, extra_job_nos,
                                                          extra_suppliers)
                               VALUES('office', ?, ?, ?, ?, ?, ?, ?)""",
                            (d, signer_id, project_id, supplier_id, po_id, extra_job_nos,
                             extra_suppliers))
            in_id = cur.lastrowid
            stats["groups_inserted"] += 1
            for ln in lines:
                existing, _bid, ok = _resolve_product_for_inbound(
                    c, ln["brand"], ln["model"], stats, ln["row_no"])
                if not ok:
                    continue
                loc_id = _get_or_create(c, "locations", "code", ln["location"]) if ln["location"] else None
                # 每行可有不同供應商
                ln_sup = (ln.get("supplier") or "").strip()
                ln_supplier_id = None
                if ln_sup:
                    primary = next(iter([p.strip() for raw in str(ln_sup).replace(",", "\n").replace("，", "\n").replace("/", "\n").splitlines() for p in [raw] if p.strip()]), None)
                    if primary:
                        ln_supplier_id = line_supplier_ids.get(primary) or _get_or_create(c, "suppliers", "name", primary)
                n = _expand_kit_or_insert_line(c, in_id, ln["brand"], ln["model"],
                                                existing, ln["qty"], loc_id, ln["serials"],
                                                stats, ln["row_no"], project_id=project_id,
                                                supplier_id=ln_supplier_id)
                stats["lines_inserted"] += n
            stats["details"].append({
                "date": d, "signer": signer, "job_no": job_no,
                "supplier": " / ".join(line_supplier_ids.keys()) if line_supplier_ids else "",
                "po_no": po_no,
                "lines": len(lines), "action": "imported",
            })
    return stats


def build_fig1_template() -> bytes:
    """產生一份只有表頭 + 一筆示範資料的 xlsx 範本。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "進貨"
    ws.append(["項目"] + FIG1_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 兩筆示範資料：同項目編號 = 同一張進貨單
    ws.append([1, "2026-06-25", "陳令佳", "所羅門股份有限公司", "AB",
               "1769-IQ32", "", 10, "蔡培君", "20260320004", "倉庫右側"])
    ws.append([1, "2026-06-25", "陳令佳", "所羅門股份有限公司", "AB",
               "1769-OB32", "", 5, "蔡培君", "20260320004", "倉庫右側"])
    widths = [6, 12, 10, 24, 12, 22, 20, 8, 10, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_fig1(file_bytes: bytes) -> list[dict]:
    """讀第一個 sheet（預設），回傳 dict 列表。空列已過濾。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_norm(x) for x in rows[0]]
    # 容錯：允許欄位順序不同，按表頭名稱對應
    idx = {h: header.index(h) if h in header else None for h in FIG1_HEADERS}
    # 供應商欄位接受 "供應商" 或舊名 "出貨對象"
    if idx["供應商"] is None:
        for alias in SUPPLIER_ALIASES:
            if alias in header:
                idx["供應商"] = header.index(alias)
                break
    # 產品名稱欄位接受別名
    if idx["產品名稱"] is None:
        for alias in MODEL_ALIASES:
            if alias in header:
                idx["產品名稱"] = header.index(alias)
                break
    item_idx = header.index("項目") if "項目" in header else None
    missing = [h for h, v in idx.items() if v is None and h in
               ("到貨日期", "品牌", "產品名稱", "進貨數量")]
    if missing:
        raise ValueError(f"缺少必要欄位：{', '.join(missing)}")

    out = []
    for i, raw in enumerate(rows[1:], start=2):  # row 2 = first data row (1-indexed)
        def cell(name):
            j = idx[name]
            return raw[j] if j is not None and j < len(raw) else None

        if all(c is None or _norm(c) == "" for c in raw):
            continue

        out.append({
            "row_no": i,
            "date": _norm(cell("到貨日期")),
            "signer": _norm(cell("簽收人")),
            "supplier": _norm(cell("供應商")),
            "brand": _norm(cell("品牌")),
            "model": _norm(cell("產品名稱")),
            "serials": _parse_serials(cell("序號")),
            "qty": _parse_qty(cell("進貨數量")),
            "requester": _norm(cell("請購人員")),
            "po_no": _norm(cell("請購PO")),
            "location": _norm(cell("存放位置")),
            "item_no": _norm(raw[item_idx]) if (item_idx is not None and item_idx < len(raw)) else "",
            "_has_item_col": item_idx is not None,
        })
    return out


def import_fig1(file_bytes: bytes, dry_run: bool = False) -> dict:
    """匯入 fig1（新竹採購進貨）。
    回傳統計：{groups, rows, lines_inserted, skipped_existing_po, errors, dry_run}"""
    rows = parse_fig1(file_bytes)
    errors = []
    # 驗證
    valid_rows = []
    for r in rows:
        msgs = []
        if not r["date"]:
            msgs.append("缺到貨日期")
        if not r["model"]:
            msgs.append("缺產品名稱")
        if r["qty"] <= 0:
            msgs.append("數量需 > 0")
        if msgs:
            errors.append({"row": r["row_no"], "msgs": msgs})
            continue
        valid_rows.append(r)

    # 分組：有「項目」欄就以項目為唯一鍵；否則退回 (date, po_no, signer, supplier)
    use_item = bool(valid_rows) and valid_rows[0].get("_has_item_col")
    groups: dict = {}
    for r in valid_rows:
        if use_item:
            key = r["item_no"] or f"__row{r['row_no']}__"
        else:
            key = (r["date"], r["po_no"], r["signer"], r["supplier"])
        g = groups.setdefault(key, {"header": {}, "lines": []})
        for k in ("date", "po_no", "signer", "supplier", "requester"):
            if not g["header"].get(k) and r.get(k):
                g["header"][k] = r[k]
        g["lines"].append(r)

    stats = {
        "total_rows": len(rows),
        "valid_rows": len(valid_rows),
        "groups": len(groups),
        "lines_inserted": 0,
        "groups_inserted": 0,
        "skipped_existing_po": 0,
        "errors": errors,
        "dry_run": dry_run,
        "details": [],
    }

    if dry_run:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            stats["details"].append({
                "date": h.get("date"), "po_no": h.get("po_no"),
                "signer": h.get("signer"), "supplier": h.get("supplier"),
                "lines": len(lines),
                "preview": [f'{ln["brand"]} {ln["model"]} x{ln["qty"]} → {ln["location"]}' for ln in lines],
            })
        return stats

    with db.tx() as c:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            d = h.get("date"); po_no = h.get("po_no")
            signer = h.get("signer"); supplier = h.get("supplier")
            # 有 PO 才檢查重複；無 PO 一律當新單寫入
            if po_no:
                existing = c.execute("""SELECT io.id FROM inbound_orders io
                                        LEFT JOIN purchase_orders po ON po.id=io.po_id
                                        WHERE po.po_no=?""", (po_no,)).fetchone()
                if existing:
                    stats["skipped_existing_po"] += 1
                    stats["details"].append({
                        "date": d, "po_no": po_no, "signer": signer, "supplier": supplier,
                        "lines": len(lines), "action": "skipped (PO 已存在)",
                    })
                    continue

            signer_id = _get_or_create(c, "staff", "name", signer, {"role": "簽收"}) if signer else None
            supplier_id = _get_or_create(c, "suppliers", "name", supplier) if supplier else None
            # 用同一個請購人員（取群組第一行）
            requester_name = next((ln["requester"] for ln in lines if ln["requester"]), "")
            requester_id = _get_or_create(c, "staff", "name", requester_name, {"role": "請購"}) if requester_name else None

            # PO 主檔（PO 為空則不建）
            po_id = None
            if po_no:
                po_row = c.execute("SELECT id FROM purchase_orders WHERE po_no=?", (po_no,)).fetchone()
                if po_row:
                    po_id = po_row["id"]
                    if requester_id:
                        c.execute("UPDATE purchase_orders SET requester_id=COALESCE(requester_id, ?) WHERE id=?",
                                  (requester_id, po_id))
                else:
                    cur = c.execute("INSERT INTO purchase_orders(po_no, date, requester_id) VALUES(?,?,?)",
                                    (po_no, d, requester_id))
                    po_id = cur.lastrowid

            cur = c.execute("""INSERT INTO inbound_orders(type, date, supplier_id, signer_id, po_id, note)
                               VALUES('hsinchu', ?, ?, ?, ?, ?)""",
                            (d, supplier_id, signer_id, po_id, None))
            in_id = cur.lastrowid
            stats["groups_inserted"] += 1

            for ln in lines:
                existing, _bid, ok = _resolve_product_for_inbound(
                    c, ln["brand"], ln["model"], stats, ln["row_no"])
                if not ok:
                    continue
                loc_id = _get_or_create(c, "locations", "code", ln["location"]) if ln["location"] else None
                n = _expand_kit_or_insert_line(c, in_id, ln["brand"], ln["model"],
                                                existing, ln["qty"], loc_id, ln["serials"],
                                                stats, ln["row_no"])
                stats["lines_inserted"] += n

            stats["details"].append({
                "date": d, "po_no": po_no, "signer": signer, "supplier": supplier,
                "lines": len(lines), "action": "imported",
            })

    return stats


# 類型欄位的值對應
TYPE_HSINCHU_ALIASES = ("新竹", "新竹採購", "hsinchu")
TYPE_OFFICE_ALIASES = ("台南辦公室", "辦公室", "辦公室請購", "office")
TYPE_SURPLUS_ALIASES = ("餘料退回", "餘料", "surplus", "surplus_return")
TYPE_BORROW_RETURN_ALIASES = ("借出歸還", "歸還", "borrow_return", "return")


def import_surplus_return(file_bytes: bytes, dry_run: bool = False) -> dict:
    """匯入餘料退回。必填：工號、料件；其餘選填。
    每筆 inbound_line 的 is_surplus 設為 1；source_outbound_line_id 留 NULL
    （Excel 無法精準對應到哪一張原始出貨單）。
    同 (date, signer, project) 視為一張單。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "valid_rows": 0, "groups": 0, "groups_inserted": 0,
                "lines_inserted": 0, "errors": [], "dry_run": dry_run, "details": []}
    header = [_norm(x) for x in rows[0]]
    model_idx = None
    for h in ("料件", "產品名稱", "型號"):
        if h in header:
            model_idx = header.index(h); break
    project_idx = None
    for h in ("工號", "對應工號"):
        if h in header:
            project_idx = header.index(h); break
    date_idx = header.index("到貨日期") if "到貨日期" in header else None
    signer_idx = header.index("簽收人") if "簽收人" in header else None
    brand_idx = header.index("品牌") if "品牌" in header else None
    qty_idx = header.index("進貨數量") if "進貨數量" in header else None
    loc_idx = header.index("存放位置") if "存放位置" in header else None
    sn_idx = header.index("序號") if "序號" in header else None
    item_idx = header.index("項目") if "項目" in header else None

    missing = []
    if model_idx is None: missing.append("料件")
    if project_idx is None: missing.append("工號")
    if missing:
        raise ValueError(f"缺少必要欄位：{', '.join(missing)}")

    errors, parsed = [], []
    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or _norm(c) == "" for c in raw):
            continue
        def cell(idx_):
            return raw[idx_] if idx_ is not None and idx_ < len(raw) else None
        d = _norm(cell(date_idx))
        signer = _norm(cell(signer_idx))
        brand = _norm(cell(brand_idx))
        model = _norm(cell(model_idx))
        job_no = _norm(cell(project_idx))
        qty = _parse_qty(cell(qty_idx))
        loc = _norm(cell(loc_idx))
        sns = _parse_serials_office(cell(sn_idx)) if sn_idx is not None else []
        item_no = _norm(cell(item_idx))
        msgs = []
        if not model: msgs.append("缺料件")
        if not job_no: msgs.append("缺工號")
        if qty <= 0: msgs.append("數量需 > 0")
        if sns and qty > 0 and len(sns) > int(qty):
            msgs.append(f"序號筆數({len(sns)})多於進貨數量({int(qty)})")
        if msgs:
            errors.append({"row": i, "msgs": msgs}); continue
        parsed.append({"row_no": i, "date": d, "signer": signer, "brand": brand,
                       "model": model, "qty": qty, "job_no": job_no,
                       "location": loc, "serials": sns, "item_no": item_no})

    groups = {}
    for r in parsed:
        if item_idx is not None:
            key = r["item_no"] or f"__row{r['row_no']}__"
        else:
            key = (r["date"], r["signer"], r["job_no"])
        g = groups.setdefault(key, {"header": {}, "lines": []})
        for k in ("date", "signer", "job_no"):
            if not g["header"].get(k) and r.get(k):
                g["header"][k] = r[k]
        g["lines"].append(r)

    stats = {
        "total_rows": len(rows) - 1, "valid_rows": len(parsed),
        "groups": len(groups), "groups_inserted": 0, "lines_inserted": 0,
        "errors": errors, "dry_run": dry_run, "details": [],
    }
    if dry_run:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            stats["details"].append({
                "date": h.get("date"), "signer": h.get("signer"),
                "job_no": h.get("job_no"), "lines": len(lines),
                "preview": [f'{ln["brand"] or ""} {ln["model"]} x{ln["qty"]} → {ln["location"]}' for ln in lines],
            })
        return stats

    with db.tx() as c:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            d = h.get("date"); signer = h.get("signer"); job_no = h.get("job_no")
            signer_id = _get_or_create(c, "staff", "name", signer, {"role": "簽收"}) if signer else None
            prow = c.execute("SELECT id FROM projects WHERE job_no=?", (job_no,)).fetchone()
            if not prow:
                cur = c.execute("INSERT INTO projects(job_no) VALUES(?)", (job_no,))
                project_id = cur.lastrowid
            else:
                project_id = prow["id"]
            cur = c.execute("""INSERT INTO inbound_orders(type, date, signer_id, project_id)
                               VALUES('surplus_return', ?, ?, ?)""", (d or None, signer_id, project_id))
            in_id = cur.lastrowid
            stats["groups_inserted"] += 1
            for ln in lines:
                existing, _bid, ok = _resolve_product_for_inbound(
                    c, ln["brand"], ln["model"], stats, ln["row_no"])
                if not ok:
                    continue
                loc_id = _get_or_create(c, "locations", "code", ln["location"]) if ln["location"] else None
                if existing and existing["is_kit"]:
                    stats["errors"].append({"row": ln["row_no"],
                        "msgs": [f"{ln['brand']} {ln['model']} 為組合件，餘料退回不展開"]})
                    continue
                prod_id = existing["id"] if existing else _get_or_create_product(
                    c, _get_or_create(c, "brands", "name", ln["brand"]) if ln["brand"] else None,
                    ln["model"], "")
                # 餘料退回：line 與 serial 一律進自由池（project_id=NULL）；
                # 退回來源仍在 inbound_orders.project_id 保留可追溯
                cur2 = c.execute("""INSERT INTO inbound_lines(inbound_id, product_id, qty, unit,
                                    location_id, is_surplus, project_id) VALUES(?,?,?,?,?,1,NULL)""",
                                 (in_id, prod_id, ln["qty"], "個", loc_id))
                line_id = cur2.lastrowid
                stats["lines_inserted"] += 1
                for sn in ln["serials"]:
                    c.execute("""INSERT OR IGNORE INTO serial_items(product_id, serial_no, status,
                                 current_location_id, inbound_line_id, is_surplus, project_id)
                                 VALUES(?,?,?,?,?,1,NULL)""",
                              (prod_id, sn, "in_stock", loc_id, line_id))
            stats["details"].append({
                "date": d, "signer": signer, "job_no": job_no,
                "lines": len(lines), "action": "imported",
            })
    return stats


def import_borrow_return(file_bytes: bytes, dry_run: bool = False) -> dict:
    """匯入借出歸還。逐筆序號比對未歸還的 borrow_records，
    結算並建立一張「借出歸還」進貨單（type=surplus_return + is_borrow_return=1）。
    必填欄位：序號；其餘（到貨日期/簽收人/存放位置）選填。
    同 (date, signer, location) 視為一張歸還單。"""
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "valid_rows": 0, "groups": 0, "groups_inserted": 0,
                "lines_inserted": 0, "serials_settled": 0, "errors": [],
                "dry_run": dry_run, "details": []}
    header = [_norm(x) for x in rows[0]]
    sn_idx = header.index("序號") if "序號" in header else None
    date_idx = header.index("到貨日期") if "到貨日期" in header else None
    signer_idx = header.index("簽收人") if "簽收人" in header else None
    loc_idx = header.index("存放位置") if "存放位置" in header else None
    brand_idx = header.index("品牌") if "品牌" in header else None
    model_idx = None
    for h in ("料件", "產品名稱", "型號"):
        if h in header:
            model_idx = header.index(h); break
    item_idx = header.index("項目") if "項目" in header else None

    if sn_idx is None:
        raise ValueError("缺少必要欄位：序號")

    errors, parsed = [], []
    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or _norm(c) == "" for c in raw):
            continue
        def cell(idx_):
            return raw[idx_] if idx_ is not None and idx_ < len(raw) else None
        d = _norm(cell(date_idx))
        signer = _norm(cell(signer_idx))
        loc = _norm(cell(loc_idx))
        brand = _norm(cell(brand_idx))
        model = _norm(cell(model_idx))
        sns = _parse_serials_office(cell(sn_idx))
        item_no = _norm(cell(item_idx))
        if not sns:
            errors.append({"row": i, "msgs": ["缺序號（借出歸還必須以序號比對）"]}); continue
        for sn in sns:
            parsed.append({"row_no": i, "date": d, "signer": signer, "location": loc,
                           "brand": brand, "model": model, "serial_no": sn,
                           "item_no": item_no})

    # 分組：同 (date, signer, location) 或 「項目」鍵 → 一張歸還單
    groups = {}
    for r in parsed:
        if item_idx is not None:
            key = r["item_no"] or f"__row{r['row_no']}__"
        else:
            key = (r["date"], r["signer"], r["location"])
        g = groups.setdefault(key, {"header": {}, "lines": []})
        for k in ("date", "signer", "location"):
            if not g["header"].get(k) and r.get(k):
                g["header"][k] = r[k]
        g["lines"].append(r)

    stats = {
        "total_rows": len(rows) - 1, "valid_rows": len(parsed),
        "groups": len(groups), "groups_inserted": 0, "lines_inserted": 0,
        "serials_settled": 0, "errors": errors, "dry_run": dry_run, "details": [],
    }

    if dry_run:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            stats["details"].append({
                "date": h.get("date"), "signer": h.get("signer"),
                "location": h.get("location"), "lines": len(lines),
                "preview": [ln["serial_no"] for ln in lines],
            })
        return stats

    with db.tx() as c:
        for key, g in groups.items():
            h = g["header"]; lines = g["lines"]
            d = h.get("date"); signer = h.get("signer"); loc = h.get("location")
            signer_id = _get_or_create(c, "staff", "name", signer, {"role": "簽收"}) if signer else None
            loc_id = _get_or_create(c, "locations", "code", loc) if loc else None

            # 預先比對 borrow_records；若所有序號都找不到，整組略過、不建單頭
            settle_map = {}  # product_id -> list of (borrow_rec_id, serial_item_id, from_project_id)
            for ln in lines:
                sn = ln["serial_no"]
                cand = c.execute("""SELECT si.id si_id, si.product_id, br.id br_id, br.from_project_id
                                    FROM serial_items si
                                    JOIN borrow_records br ON br.serial_item_id = si.id
                                    WHERE si.serial_no=? AND br.returned_at IS NULL
                                    ORDER BY br.borrowed_at DESC LIMIT 1""", (sn,)).fetchone()
                if not cand:
                    stats["errors"].append({"row": ln["row_no"],
                        "msgs": [f"序號 {sn} 找不到「未歸還」的借出紀錄"]})
                    continue
                settle_map.setdefault(cand["product_id"], []).append({
                    "br_id": cand["br_id"], "si_id": cand["si_id"],
                    "from_project_id": cand["from_project_id"], "serial_no": sn,
                })
            if not settle_map:
                stats["details"].append({
                    "date": d, "signer": signer, "location": loc,
                    "lines": len(lines), "action": "skipped (無可結算序號)",
                })
                continue

            cur = c.execute("""INSERT INTO inbound_orders(type, date, signer_id, is_borrow_return)
                               VALUES('surplus_return', ?, ?, 1)""", (d or None, signer_id))
            in_id = cur.lastrowid
            stats["groups_inserted"] += 1

            for pid, recs in settle_map.items():
                cur2 = c.execute("""INSERT INTO inbound_lines
                                    (inbound_id, product_id, qty, unit, location_id,
                                     is_surplus, project_id) VALUES(?,?,?,?,?,0,NULL)""",
                                 (in_id, pid, float(len(recs)), "個", loc_id))
                line_id = cur2.lastrowid
                stats["lines_inserted"] += 1
                for r in recs:
                    c.execute("""UPDATE serial_items
                                 SET status='in_stock', current_location_id=?, project_id=?,
                                     inbound_line_id=?, outbound_line_id=NULL, is_surplus=0
                                 WHERE id=?""",
                              (loc_id, r["from_project_id"], line_id, r["si_id"]))
                    c.execute("""UPDATE borrow_records
                                 SET returned_at=?, returned_inbound_line_id=?
                                 WHERE id=?""", (d or None, line_id, r["br_id"]))
                    stats["serials_settled"] += 1
            stats["details"].append({
                "date": d, "signer": signer, "location": loc,
                "lines": len(lines), "settled": stats["serials_settled"],
                "action": "imported",
            })
    return stats


def import_inbound_auto(file_bytes: bytes, dry_run: bool = False) -> dict:
    """讀 xlsx 的「類型」欄，自動分流到 import_fig1 / import_office 後合併結果。
    若 Excel 沒有「類型」欄，整檔當作新竹採購（fig1）處理。
    """
    from io import BytesIO
    from openpyxl import Workbook
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel 內沒有任何 sheet")
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 是空的")
    header = [_norm(x) for x in rows[0]]
    if "類型" not in header:
        return import_fig1(file_bytes, dry_run=dry_run)

    type_idx = header.index("類型")
    keep_cols = [i for i, h in enumerate(header) if h != "類型"]
    out_header = [header[i] for i in keep_cols]

    def build_subset(subset_rows):
        nb = Workbook(); nws = nb.active; nws.title = "進貨"
        nws.append(out_header)
        for r in subset_rows:
            nws.append([r[i] if i < len(r) else None for i in keep_cols])
        buf = BytesIO(); nb.save(buf); return buf.getvalue()

    hs_rows, of_rows, sr_rows, br_rows, unknown_rows = [], [], [], [], []
    for r in rows[1:]:
        if all(c is None or _norm(c) == "" for c in r):
            continue
        t = _norm(r[type_idx]).strip() if type_idx < len(r) else ""
        if t in TYPE_HSINCHU_ALIASES:
            hs_rows.append(r)
        elif t in TYPE_OFFICE_ALIASES:
            of_rows.append(r)
        elif t in TYPE_SURPLUS_ALIASES:
            sr_rows.append(r)
        elif t in TYPE_BORROW_RETURN_ALIASES:
            br_rows.append(r)
        else:
            unknown_rows.append((r, t))

    merged = {
        "total_rows": len(rows) - 1,
        "valid_rows": 0,
        "groups": 0, "groups_inserted": 0, "lines_inserted": 0,
        "skipped_existing_po": 0,
        "errors": [], "dry_run": dry_run, "details": [],
        "by_type": {},
    }

    if hs_rows:
        r1 = import_fig1(build_subset(hs_rows), dry_run=dry_run)
        merged["by_type"]["hsinchu"] = {
            "total": r1.get("total_rows", 0),
            "groups_inserted": r1.get("groups_inserted", 0),
            "lines_inserted": r1.get("lines_inserted", 0),
        }
        merged["valid_rows"] += r1.get("valid_rows", 0)
        merged["groups"] += r1.get("groups", 0)
        merged["groups_inserted"] += r1.get("groups_inserted", 0)
        merged["lines_inserted"] += r1.get("lines_inserted", 0)
        merged["skipped_existing_po"] += r1.get("skipped_existing_po", 0)
        for e in r1.get("errors", []):
            merged["errors"].append({**e, "section": "新竹"})
        for d in r1.get("details", []):
            merged["details"].append({**d, "type_label": "新竹", "section": "hsinchu"})

    if of_rows:
        r2 = import_office(build_subset(of_rows), dry_run=dry_run)
        merged["by_type"]["office"] = {
            "total": r2.get("total_rows", 0),
            "groups_inserted": r2.get("groups_inserted", 0),
            "lines_inserted": r2.get("lines_inserted", 0),
        }
        merged["valid_rows"] += r2.get("valid_rows", 0)
        merged["groups"] += r2.get("groups", 0)
        merged["groups_inserted"] += r2.get("groups_inserted", 0)
        merged["lines_inserted"] += r2.get("lines_inserted", 0)
        for e in r2.get("errors", []):
            merged["errors"].append({**e, "section": "辦公室"})
        for d in r2.get("details", []):
            merged["details"].append({**d, "type_label": "台南辦公室", "section": "office"})

    if sr_rows:
        r3 = import_surplus_return(build_subset(sr_rows), dry_run=dry_run)
        merged["by_type"]["surplus_return"] = {
            "total": r3.get("total_rows", 0),
            "groups_inserted": r3.get("groups_inserted", 0),
            "lines_inserted": r3.get("lines_inserted", 0),
        }
        merged["valid_rows"] += r3.get("valid_rows", 0)
        merged["groups"] += r3.get("groups", 0)
        merged["groups_inserted"] += r3.get("groups_inserted", 0)
        merged["lines_inserted"] += r3.get("lines_inserted", 0)
        for e in r3.get("errors", []):
            merged["errors"].append({**e, "section": "餘料退回"})
        for d in r3.get("details", []):
            merged["details"].append({**d, "type_label": "餘料退回", "section": "surplus_return"})

    if br_rows:
        r4 = import_borrow_return(build_subset(br_rows), dry_run=dry_run)
        merged["by_type"]["borrow_return"] = {
            "total": r4.get("total_rows", 0),
            "groups_inserted": r4.get("groups_inserted", 0),
            "lines_inserted": r4.get("lines_inserted", 0),
            "serials_settled": r4.get("serials_settled", 0),
        }
        merged["valid_rows"] += r4.get("valid_rows", 0)
        merged["groups"] += r4.get("groups", 0)
        merged["groups_inserted"] += r4.get("groups_inserted", 0)
        merged["lines_inserted"] += r4.get("lines_inserted", 0)
        for e in r4.get("errors", []):
            merged["errors"].append({**e, "section": "借出歸還"})
        for d in r4.get("details", []):
            merged["details"].append({**d, "type_label": "借出歸還", "section": "borrow_return"})

    for r, t in unknown_rows:
        merged["errors"].append({"row": "?", "msgs": [f"未知的「類型」值：{t!r}（僅接受 新竹／台南辦公室／餘料退回／借出歸還）"]})

    return merged


# ---------- Raw 匯入：完整保留欄位，不解析業務語意 ----------
RAW_HEADER_MAP = {
    "ITEM": "item_no",
    "到貨日期": "date", "進貨日期": "date", "日期": "date",
    "簽收人": "signer",
    "出貨對象": "source", "供應商": "source",
    "商品名稱": "model", "品牌": "model", "產品名稱": "model", "料件": "model", "型號": "model",
    "商品敘述": "description", "敘述": "description",
    "序號": "serial_no", "SN": "serial_no",
    "數量": "qty", "進貨數量": "qty",
    "完成工號": "project_no", "工號": "project_no", "對應工號": "project_no",
    "完成案主": "owner", "案主": "owner",
    "案名": "project_name",
    "備註": "note",
    "庫存料件": "stock_item",
    "庫存數量": "stock_qty",
    "位置": "location", "存放位置": "location",
    "取放人": "picker", "請購人員": "picker",
    "領出借出單號": "ledger_no", "領出\n借出單號": "ledger_no", "借出單號": "ledger_no",
    "進貨單頁次": "page_no", "頁次": "page_no",
    "編號位置": "code_pos",
    "color": "color", "顏色": "color",
    "PO": "po_no", "po": "po_no", "PO號": "po_no", "請購PO": "po_no",
    "回傳": "photo_sent", "回傳總部": "photo_sent",
    # 新版 test_raw 欄位別名
    "對應業主": "owner", "完成案主": "owner",
    "對應工號": "project_no",
    "產品名稱": "model", "產品敘述": "description",
    "存放位置": "location",
    "交貨單頁數": "page_no", "頁數": "page_no",
    "勁傑借出單號": "ledger_no", "勁傑\n借出單號": "ledger_no",
    "轉交": "picker", "轉交日期": None,  # 轉交日期目前不對應任何欄位
}


def import_raw_excel(file_bytes: bytes, batch_id: str) -> dict:
    """讀取 raw excel → 插入 raw_imports 暫存表（status='pending'）。
    序號欄寫入前必經 db.normalize_sn（憲法第三條）。
    """
    import io
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"rows_inserted": 0, "total_rows": 0,
                "headers_matched": [], "headers_unknown": [], "errors": []}

    headers = list(rows[0])
    body = rows[1:]
    col_map = {}
    unknown = []
    matched = []
    for idx, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        key_no_nl = key.replace("\n", "").replace(" ", "")
        target = RAW_HEADER_MAP.get(key) or RAW_HEADER_MAP.get(key_no_nl)
        if target:
            col_map[idx] = target
            matched.append((str(h), target))
        else:
            unknown.append(str(h))

    inserted = 0
    errors = []
    with db.tx() as c:
        for r_no, row in enumerate(body, start=2):
            data = {col_map[i]: row[i] for i in range(len(row))
                    if i in col_map and row[i] not in (None, "")}
            if not data:
                continue
            for dcol in ("date", "photo_sent"):
                if dcol in data:
                    v = data[dcol]
                    if isinstance(v, datetime):
                        data[dcol] = v.date().isoformat()
                    elif isinstance(v, date):
                        data[dcol] = v.isoformat()
                    else:
                        data[dcol] = str(v).strip() if v else None
            if "qty" in data:
                try:
                    data["qty"] = float(data["qty"])
                except (TypeError, ValueError):
                    data["qty"] = None
            if "stock_qty" in data:
                try:
                    data["stock_qty"] = float(data["stock_qty"])
                except (TypeError, ValueError):
                    data["stock_qty"] = None
            if "serial_no" in data:
                raw_sn = str(data["serial_no"]).strip()
                if raw_sn:
                    tokens = []
                    for tok in raw_sn.replace("\n", "/").replace(",", "/").replace("，", "/").split("/"):
                        n = db.normalize_sn(tok)
                        if n:
                            tokens.append(n)
                    data["serial_no"] = "/".join(tokens) if tokens else None
                else:
                    data["serial_no"] = None
            for k, v in list(data.items()):
                if k in ("qty", "stock_qty"):
                    continue
                if v is not None:
                    if not isinstance(v, str):
                        v = str(v)
                    v = v.strip()
                    data[k] = v if v else None

            data["batch_id"] = batch_id
            data["source_row"] = r_no
            data["status"] = "pending"

            # 若 description 缺、model 在 products 主檔有對應 → 自動補
            if not data.get("description") and data.get("model"):
                pr = c.execute("""SELECT description FROM products
                                  WHERE model=? AND description IS NOT NULL
                                    AND description <> '' LIMIT 1""",
                               (data["model"],)).fetchone()
                if pr:
                    data["description"] = pr["description"]

            cols = list(data.keys())
            placeholders = ",".join("?" * len(cols))
            try:
                c.execute(f"INSERT INTO raw_imports({','.join(cols)}) VALUES({placeholders})",
                          [data[k] for k in cols])
                inserted += 1
            except Exception as e:
                errors.append({"row": r_no, "msgs": [str(e)]})

    return {
        "rows_inserted": inserted,
        "total_rows": len(body),
        "headers_matched": matched,
        "headers_unknown": unknown,
        "errors": errors,
    }

